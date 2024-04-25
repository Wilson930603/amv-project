# Define your item pipelines here
#
# Don't forget to add your pipeline to the ITEM_PIPELINES setting
# See: https://docs.scrapy.org/en/latest/topics/item-pipeline.html


# useful for handling different item types with a single interface
from openpyxl import Workbook
from openpyxl.styles import Font
import os,re,html

class GetdataPipeline:
    def open_spider(self,spider):
        if not os.path.exists('./Data'):
            os.mkdir('./Data',0o777)
        self.wb=Workbook()
        self.SHEET=[]
        self.DATA=[]
        self.FIELDS={}
    def close_spider(self,spider):
        wskpi=self.wb.active
        for item in self.DATA:
            if len(self.SHEET)==0:
                wskpi.title=item['SHEET']        
                self.SHEET.append(item['SHEET'])
                TT=0
                for key in self.FIELDS[item['SHEET']]:
                    header_font = Font(bold=True)
                    O=wskpi.cell(column=TT+1, row=1, value=key)
                    O.font = header_font
                    TT+=1
            elif not item['SHEET'] in self.SHEET:
                wskpi=self.wb.create_sheet()                
                wskpi.title=item['SHEET']
                self.SHEET.append(item['SHEET'])
                TT=0
                for key in self.FIELDS[item['SHEET']]:
                    header_font = Font(bold=True)
                    O=wskpi.cell(column=TT+1, row=1, value=key)
                    O.font = header_font
                    TT+=1
            else:
                wskpi=self.wb[item['SHEET']]
            Data=[]
            for key in self.FIELDS[item['SHEET']]:
                if key in item:
                    VALUE=item[key]
                else:
                    VALUE=''
                if str(VALUE).lower()=='none':
                    VALUE=''
                TYPE=self.get_DataType(VALUE)
                if TYPE=='INT':
                    Data.append(int(VALUE))
                elif TYPE=='FLOAT':
                    Data.append(float(VALUE))
                else:
                    Data.append(html.unescape(VALUE))
                
            wskpi.append(Data)
        for i in range(len(self.SHEET)):
            ws=self.wb[self.SHEET[i]]
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                if length>100:
                    length=100
                ws.column_dimensions[column_cells[0].column_letter].width = length+2      
        self.wb.save('./Data/Data_'+str(spider.name)+'.xlsx')
    def process_item(self, item, spider):
        if len(item)>0:
            if not 'SHEET' in item:
                item['SHEET']=spider.name
            if not item['SHEET'] in self.FIELDS:
                self.FIELDS[item['SHEET']]=[]
            for key in item.keys():
                if not key in self.FIELDS[item['SHEET']] and key!='SHEET':
                    self.FIELDS[item['SHEET']].append(key)
            self.DATA.append(item)
        return item
    def Get_Number(self,xau):
        KQ=re.sub(r"([^0-9.-])","", str(xau).strip())
        return KQ
    def get_DataType(self,strtxt):
        strtxt=str(strtxt).strip()
        if len(strtxt) == 0: return 'BLANK'
        if re.match(r'([-+]\s*)?\d+[lL]?$', strtxt): 
            return 'INT'
        if re.match(r'([-+]\s*)?[1-9][0-9]*\.?[0-9]*([Ee][+-]?[0-9]+)?$', strtxt): 
            return 'FLOAT'
        if re.match(r'([-+]\s*)?[0-9]*\.?[0-9][0-9]*([Ee][+-]?[0-9]+)?$', strtxt): 
            return 'FLOAT'
        return 'TEXT'